from pathlib import Path

import openpyxl

from marktview.excel_writer import (
    HEADERS,
    load_existing_listing_ids,
    write_listings_to_excel,
)
from marktview.models import Listing


def test_load_existing_listing_ids_empty(tmp_path: Path):
    excel_path = tmp_path / "missing.xlsx"
    assert load_existing_listing_ids(excel_path) == set()

    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.append(HEADERS)
    worksheet.append([None])
    workbook.save(excel_path)

    assert load_existing_listing_ids(excel_path) == set()


def test_write_and_load_listing_ids(tmp_path: Path):
    excel_path = tmp_path / "anzeigen.xlsx"
    listings = [
        Listing(title="A", url="https://a", listing_id="123"),
        Listing(title="B", url="https://b", listing_id="456"),
    ]

    output_path = write_listings_to_excel(listings, excel_path)
    assert output_path == excel_path

    workbook = openpyxl.load_workbook(excel_path)
    worksheet = workbook.active

    assert worksheet["A1"].value == HEADERS[0]
    assert worksheet.max_row == 3

    stored_ids = load_existing_listing_ids(excel_path)
    assert stored_ids == {"123", "456"}

    # Duplicate should be skipped when writing again
    write_listings_to_excel([Listing(title="A2", url="https://a2", listing_id="123")], excel_path)
    workbook = openpyxl.load_workbook(excel_path)
    worksheet = workbook.active
    assert worksheet.max_row == 3
