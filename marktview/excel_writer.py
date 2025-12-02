"""Excel export helpers for Markt.de listings."""

from pathlib import Path
from typing import Iterable, Set

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment

from .models import Listing


HEADERS = [
    "Titel",
    "Url",
    "PLZ",
    "Erstellungsdatum",
    "Text",
    "Geschlecht",
    "Zielgruppe",
    "Geldinteresse",
    "Anzeigenkennung",
    "Benutzername",
]


def load_existing_listing_ids(path: str | Path) -> Set[str]:
    """Return listing IDs already stored in the Excel file (if any)."""

    excel_path = Path(path)
    if not excel_path.exists():
        return set()

    workbook = load_workbook(excel_path)
    worksheet = workbook.active

    listing_ids: Set[str] = set()
    for row in worksheet.iter_rows(min_row=2, values_only=True):
        if not row or len(row) < 9:  # pragma: no cover - defensive guard
            continue
        listing_id = row[8]
        if listing_id:
            listing_ids.add(str(listing_id))

    return listing_ids


def write_listings_to_excel(listings: Iterable[Listing], path: str | Path) -> Path:
    """Write the provided listings to an Excel file."""

    output_path = Path(path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    existing_ids = load_existing_listing_ids(output_path)

    if output_path.exists():
        workbook = load_workbook(output_path)
        worksheet = workbook.active
    else:
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Anzeigen"
        worksheet.append(HEADERS)
        worksheet.column_dimensions["A"].width = 40
        worksheet.column_dimensions["B"].width = 80
        worksheet.column_dimensions["C"].width = 10
        worksheet.column_dimensions["D"].width = 20
        worksheet.column_dimensions["E"].width = 100
        worksheet.column_dimensions["F"].width = 20
        worksheet.column_dimensions["G"].width = 20
        worksheet.column_dimensions["H"].width = 20
        worksheet.column_dimensions["I"].width = 20
        worksheet.column_dimensions["J"].width = 30

    for listing in listings:
        if listing.listing_id and listing.listing_id in existing_ids:
            continue

        worksheet.append(
            [
                listing.title,
                listing.url,
                listing.postal_code,
                listing.created_at,
                listing.body,
                listing.gender,
                listing.target_audience,
                listing.financial_interest,
                listing.listing_id,
                listing.username,
            ]
        )

        if listing.listing_id:
            existing_ids.add(listing.listing_id)

    wrap_alignment = Alignment(wrap_text=True)
    for cell in worksheet["E:E"]:
        cell.alignment = wrap_alignment

    workbook.save(output_path)
    print(f"[SUCCESS] Excel-Datei gespeichert unter: {output_path}")
    return output_path
